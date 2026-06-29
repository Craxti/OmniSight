import time

from fastapi.testclient import TestClient

from tests.rsm_helpers import seed_chain as _seed_chain
from tests.seed_helpers import create_ci, get_type_map
from tests.v1_helpers import (
    API_V1,
    audit_items,
    ci,
    dashboard,
    export_payload,
    items,
    report,
    search,
    validation,
)


def test_login_and_ci_crud(client: TestClient, auth_headers: dict):
    r = client.post(
        f"{API_V1}/ci",
        json={"name": "web-01", "type_name": "Server", "owner": "ops", "environment": "test", "criticality": "medium"},
        headers=auth_headers,
    )
    assert r.status_code == 200
    ci_id = ci(r.json())["id"]

    r = client.get(f"{API_V1}/ci/{ci_id}", headers=auth_headers)
    assert r.status_code == 200
    assert ci(r.json())["name"] == "web-01"

    r = client.patch(f"{API_V1}/ci/{ci_id}", json={"status": "temporarily_disabled"}, headers=auth_headers)
    assert r.status_code == 200
    assert ci(r.json())["status"] == "temporarily_disabled"


def test_correlation_ingest_four_alerts(client: TestClient, auth_headers: dict):
    _seed_chain(client, auth_headers)
    alerts = [
        {"hostname": "app-01"},
        {"ip": "10.0.0.5"},
        {"externalId": "ext-db-1"},
        {"serviceCode": "PAY", "applicationCode": "PAY-APP"},
    ]
    r = client.post(f"{API_V1}/correlation/ingest", json={"alerts": alerts, "source": "pytest"}, headers=auth_headers)
    assert r.status_code == 200
    data = r.json()
    assert data["schema_version"] == "rsm-correlation-v1"
    assert len(data["resolve"]["resolved"]) == 4
    assert data["correlation"]["chain_related"] is True


def test_export_filtered(client: TestClient, auth_headers: dict):
    _seed_chain(client, auth_headers)
    r = client.get(f"{API_V1}/ci/export/full?environment=production", headers=auth_headers)
    assert r.status_code == 200
    payload = export_payload(r.json())
    assert len(payload["elements"]) >= 3
    assert all(e.get("environment") == "production" for e in payload["elements"])
    assert len(payload["relations"]) >= 2


def test_export_xlsx_includes_elements_and_relations(client: TestClient, auth_headers: dict):
    _seed_chain(client, auth_headers)
    r = client.get(f"{API_V1}/ci/export/xlsx?environment=production", headers=auth_headers)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("content-type", "")
    from io import BytesIO

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(r.content))
    assert "Elements" in wb.sheetnames
    assert "Relations" in wb.sheetnames
    assert wb["Elements"].max_row >= 4
    assert wb["Relations"].max_row >= 3


def test_export_csv_zip_includes_relations(client: TestClient, auth_headers: dict):
    import json
    import zipfile
    from io import BytesIO

    _seed_chain(client, auth_headers)
    r = client.get(f"{API_V1}/ci/export/csv", headers=auth_headers)
    assert r.status_code == 200
    assert r.headers.get("content-type", "").startswith("application/zip")
    with zipfile.ZipFile(BytesIO(r.content)) as zf:
        names = set(zf.namelist())
        assert "elements.csv" in names
        assert "relations.csv" in names
        manifest = json.loads(zf.read("manifest.json"))
        assert manifest["relation_count"] >= 2


def test_validate_relations_detects_cycle(client: TestClient, auth_headers: dict):
    from src.core import database as db_module
    from src.models import Relation

    types = get_type_map(client, auth_headers)

    def mk(name: str):
        return create_ci(client, auth_headers, types, name, "Server")

    a, b, c = mk("cycle-a"), mk("cycle-b"), mk("cycle-c")
    for src, tgt in [(a, b), (b, c)]:
        r = client.post(
            f"{API_V1}/relations",
            json={"source_ci_id": src, "target_ci_id": tgt, "relation_type": "depends_on", "status": "active"},
            headers=auth_headers,
        )
        assert r.status_code == 200, r.text

    db = db_module.SessionLocal()
    try:
        db.add(Relation(source_ci_id=c, target_ci_id=a, relation_type="depends_on", status="active"))
        db.commit()
    finally:
        db.close()

    r = client.get(f"{API_V1}/relations/validate", headers=auth_headers)
    assert r.status_code == 200
    body = validation(r.json())
    assert body["valid"] is False
    assert any(i["type"] == "cycle" for i in body["issues"])


def test_create_relation_blocks_depends_on_cycle(client: TestClient, auth_headers: dict):
    types = get_type_map(client, auth_headers)

    def mk(name: str):
        return create_ci(client, auth_headers, types, name, "Server")

    a, b, c = mk("block-cycle-a"), mk("block-cycle-b"), mk("block-cycle-c")
    for src, tgt in [(a, b), (b, c)]:
        r = client.post(
            f"{API_V1}/relations",
            json={"source_ci_id": src, "target_ci_id": tgt, "relation_type": "depends_on", "status": "active"},
            headers=auth_headers,
        )
        assert r.status_code == 200, r.text
    r = client.post(
        f"{API_V1}/relations",
        json={"source_ci_id": c, "target_ci_id": a, "relation_type": "depends_on", "status": "active"},
        headers=auth_headers,
    )
    assert r.status_code == 400


def test_import_relations_json(client: TestClient, auth_headers: dict):
    db_id, app_id, _ = _seed_chain(client, auth_headers)
    r = client.post(
        f"{API_V1}/relations/import/json",
        json={
            "relations": [
                {
                    "source_ci_id": app_id,
                    "target_ci_id": db_id,
                    "relation_type": "uses",
                    "status": "active",
                    "data_source": "pytest",
                }
            ]
        },
        headers=auth_headers,
    )
    assert r.status_code == 200
    assert report(r.json())["created"] == 1

    r = client.get(f"{API_V1}/relations", headers=auth_headers)
    assert any(rel["relation_type"] == "uses" and rel["source_ci_id"] == app_id for rel in items(r.json()))


def test_health_includes_database(client: TestClient):
    r = client.get("/health")
    assert r.status_code == 200
    body = r.json()
    assert body["status"] == "ok"
    assert body["database"] is True
    assert body["nfr"]["fr_coverage"] == "1-54"
    assert body["nfr"]["resolve_search_target_ms"] == 200


def test_dashboard_overview_includes_health_and_audit(client: TestClient, auth_headers: dict):
    r = client.get(f"{API_V1}/dashboard/overview", headers=auth_headers)
    assert r.status_code == 200
    body = dashboard(r.json())
    assert "model_health" in body
    assert "valid" in body["model_health"]
    assert "recent_audit" in body


def test_nfr_health_reports_nfr_metadata(client: TestClient):
    r = client.get("/health/ready")
    assert r.status_code == 200
    nfr = r.json().get("nfr", {})
    assert nfr.get("resolve_search_target_ms") == 200
    assert "rate_limit" in nfr
    assert "log_json" in nfr


NFR_MAX_MS = 200


def test_nfr_resolve_and_search_under_200ms(client: TestClient, auth_headers: dict):
    """§9 NFR: search/resolve suitable for external correlation (local benchmark)."""
    _seed_chain(client, auth_headers)
    alerts = [{"hostname": "app-01"}, {"ip": "10.0.0.5"}]

    t0 = time.perf_counter()
    r = client.post(f"{API_V1}/resources/resolve", json={"alerts": alerts}, headers=auth_headers)
    resolve_ms = (time.perf_counter() - t0) * 1000
    assert r.status_code == 200, r.text
    assert resolve_ms < NFR_MAX_MS, f"resolve took {resolve_ms:.1f}ms (limit {NFR_MAX_MS}ms)"

    t0 = time.perf_counter()
    r = client.get(f"{API_V1}/resources/search?hostname=app-01", headers=auth_headers)
    search_ms = (time.perf_counter() - t0) * 1000
    assert r.status_code == 200, r.text
    assert search_ms < NFR_MAX_MS, f"search took {search_ms:.1f}ms (limit {NFR_MAX_MS}ms)"
    assert search(r.json())["total"] >= 1


def test_duplicate_ci_name_rejected(client: TestClient, auth_headers: dict):
    body = {"name": "unique-ci", "type_name": "Server", "owner": "ops", "environment": "test", "criticality": "low"}
    r = client.post(f"{API_V1}/ci", json=body, headers=auth_headers)
    assert r.status_code == 200, r.text
    r = client.post(f"{API_V1}/ci", json=body, headers=auth_headers)
    assert r.status_code == 409


def test_ci_import_skips_unchanged_rows(client: TestClient, auth_headers: dict):
    db_id, _, _ = _seed_chain(client, auth_headers)
    r = client.get(f"{API_V1}/ci/{db_id}", headers=auth_headers)
    snapshot = ci(r.json())
    items_payload = [
        {
            "name": snapshot["name"],
            "type_name": "Database",
            "status": snapshot["status"],
            "criticality": snapshot["criticality"],
            "environment": snapshot["environment"],
            "owner": snapshot["owner"],
            "team": snapshot.get("team"),
            "attributes": snapshot.get("attributes") or {},
            "external_ids": snapshot.get("external_ids") or {},
        }
    ]
    r = client.post(f"{API_V1}/ci/import", json=items_payload, headers=auth_headers)
    assert r.status_code == 200, r.text
    body = report(r.json())
    assert body["skipped"] == 1
    assert body["updated"] == 0
    assert body["created"] == 0


def test_ci_list_filters_by_hostname_and_external_id(client: TestClient, auth_headers: dict):
    db_id, _, _ = _seed_chain(client, auth_headers)
    r = client.get(f"{API_V1}/ci?hostname=app-01", headers=auth_headers)
    assert r.status_code == 200
    names = {c["name"] for c in items(r.json())}
    assert "demo-app" in names

    r = client.get(f"{API_V1}/ci?external_id=ext-db-1", headers=auth_headers)
    assert r.status_code == 200
    assert any(c["id"] == db_id for c in items(r.json()))


def test_resources_search_by_type_name(client: TestClient, auth_headers: dict):
    _seed_chain(client, auth_headers)
    r = client.get(f"{API_V1}/resources/search?type_name=Business%20Service", headers=auth_headers)
    assert r.status_code == 200
    sbody = search(r.json())
    assert sbody["total"] >= 1
    assert all(i["type"] == "Business Service" for i in sbody["items"])


def test_export_by_business_service_subtree(client: TestClient, auth_headers: dict):
    _, _, biz_id = _seed_chain(client, auth_headers)
    r = client.get(f"{API_V1}/ci/export/full?business_service_id={biz_id}", headers=auth_headers)
    assert r.status_code == 200
    payload = export_payload(r.json())
    names = {e["name"] for e in payload["elements"]}
    assert {"demo-biz", "demo-app", "demo-db"}.issubset(names)


def test_duplicate_relation_allowed(client: TestClient, auth_headers: dict):
    types = get_type_map(client, auth_headers)

    def mk(name: str):
        return create_ci(client, auth_headers, types, name, "Server")

    a, b = mk("rel-a"), mk("rel-b")
    rel = {"source_ci_id": a, "target_ci_id": b, "relation_type": "depends_on"}
    r = client.post(f"{API_V1}/relations", json=rel, headers=auth_headers)
    assert r.status_code == 200, r.text
    r = client.post(f"{API_V1}/relations", json=rel, headers=auth_headers)
    assert r.status_code == 200, r.text


def test_resources_search_by_owner(client: TestClient, auth_headers: dict):
    _seed_chain(client, auth_headers)
    r = client.get(f"{API_V1}/resources/search?owner=ops", headers=auth_headers)
    assert r.status_code == 200, r.text
    assert search(r.json())["total"] >= 1


def test_resources_detail_and_relations(client: TestClient, auth_headers: dict):
    db_id, app_id, _ = _seed_chain(client, auth_headers)
    r = client.get(f"{API_V1}/resources/{db_id}", headers=auth_headers)
    assert r.status_code == 200, r.text
    assert ci(r.json())["name"] == "demo-db"

    r = client.get(f"{API_V1}/resources/{db_id}/relations", headers=auth_headers)
    assert r.status_code == 200, r.text
    rel_items = items(r.json())
    assert any(rel["source_ci_id"] == app_id and rel["target_ci_id"] == db_id for rel in rel_items)


def test_correlation_chain_check(client: TestClient, auth_headers: dict):
    db_id, app_id, biz_id = _seed_chain(client, auth_headers)
    r = client.post(
        f"{API_V1}/correlation/chain-check",
        json={"resource_ids": [db_id, app_id, biz_id]},
        headers=auth_headers,
    )
    assert r.status_code == 200, r.text
    assert r.json()["chain_related"] is True


def test_relation_delete_archives_status(client: TestClient, auth_headers: dict):
    db_id, app_id, _ = _seed_chain(client, auth_headers)
    r = client.get(f"{API_V1}/relations", headers=auth_headers)
    rel = next(rel for rel in items(r.json()) if rel["source_ci_id"] == app_id and rel["target_ci_id"] == db_id)
    r = client.delete(f"{API_V1}/relations/{rel['id']}", headers=auth_headers)
    assert r.status_code == 200, r.text
    r = client.get(f"{API_V1}/audit/relation/{rel['id']}", headers=auth_headers)
    assert r.status_code == 200
    deleted_entry = next((e for e in audit_items(r.json()) if e["action"] == "delete"), None)
    assert deleted_entry is not None
    assert deleted_entry["new_value"]["status"] == "archived"


def test_ci_audit_includes_relation_changes(client: TestClient, auth_headers: dict):
    db_id, app_id, _ = _seed_chain(client, auth_headers)
    r = client.patch(f"{API_V1}/ci/{db_id}", json={"description": "history check"}, headers=auth_headers)
    assert r.status_code == 200, r.text
    r = client.get(f"{API_V1}/audit/ci/{db_id}", headers=auth_headers)
    assert r.status_code == 200, r.text
    entries = audit_items(r.json())
    assert any(e["entity_type"] == "ci" and e["action"] == "update" for e in entries)
    rel_entry = next(
        (
            e
            for e in entries
            if e["entity_type"] == "relation"
            and e["action"] == "create"
            and e["new_value"]["source_ci_id"] == app_id
            and e["new_value"]["target_ci_id"] == db_id
        ),
        None,
    )
    assert rel_entry is not None


def test_import_relation_rejects_depends_on_cycle(client: TestClient, auth_headers: dict):
    types = get_type_map(client, auth_headers)

    def mk(name: str):
        return create_ci(client, auth_headers, types, name, "Server")

    a, b, c = mk("imp-cycle-a"), mk("imp-cycle-b"), mk("imp-cycle-c")
    body = report(
        client.post(
            f"{API_V1}/relations/import/json",
            json={
                "relations": [
                    {"source_ci_id": a, "target_ci_id": b, "relation_type": "depends_on", "status": "active"},
                    {"source_ci_id": b, "target_ci_id": c, "relation_type": "depends_on", "status": "active"},
                    {"source_ci_id": c, "target_ci_id": a, "relation_type": "depends_on", "status": "active"},
                ]
            },
            headers=auth_headers,
        ).json()
    )
    assert body["created"] == 2
    assert any("cycle" in err for err in body["errors"])


def test_ci_import_preview_unknown_types(client: TestClient, auth_headers: dict):
    items_payload = [
        {"name": "preview-edge-1", "type_name": "Edge Gateway Preview", "status": "active"},
        {"name": "preview-edge-2", "type_name": "Edge Gateway Preview", "status": "active"},
        {"name": "preview-srv", "type_name": "Server", "status": "active"},
    ]
    r = client.post(f"{API_V1}/ci/import/preview", json=items_payload, headers=auth_headers)
    assert r.status_code == 200, r.text
    body = export_payload(r.json())
    assert body["needs_mapping"] is True
    by_source = {p["source_type"]: p for p in body["proposals"]}
    assert by_source["Server"]["status"] == "matched"
    assert by_source["Edge Gateway Preview"]["status"] == "unknown"
    assert by_source["Edge Gateway Preview"]["item_count"] == 2
    assert by_source["Edge Gateway Preview"]["draft"]["name"] == "Edge Gateway Preview"


def test_ci_import_mapped_creates_types_and_elements(client: TestClient, auth_headers: dict):
    suffix = str(int(time.time() * 1000))
    type_a = f"E2E Stream Node {suffix}"
    type_b = f"E2E Data Vault {suffix}"
    ci_names = [f"e2e-mapped-{suffix}-a", f"e2e-mapped-{suffix}-b", f"e2e-mapped-{suffix}-c"]
    items_payload = [
        {
            "name": ci_names[0],
            "type_name": type_a,
            "status": "active",
            "criticality": "medium",
            "environment": "test",
            "owner": "e2e",
            "attributes": {"region": "eu-1"},
        },
        {
            "name": ci_names[1],
            "type_name": type_a,
            "status": "active",
            "criticality": "low",
            "environment": "test",
            "owner": "e2e",
            "attributes": {"region": "eu-2"},
        },
        {
            "name": ci_names[2],
            "type_name": type_b,
            "status": "active",
            "criticality": "high",
            "environment": "staging",
            "owner": "e2e",
            "attributes": {"bucket": "raw"},
        },
    ]
    preview = client.post(f"{API_V1}/ci/import/preview", json=items_payload, headers=auth_headers)
    assert preview.status_code == 200, preview.text
    assert export_payload(preview.json())["needs_mapping"] is True

    mappings = []
    for proposal in export_payload(preview.json())["proposals"]:
        if proposal["status"] == "matched":
            continue
        mappings.append(
            {
                "source_type": proposal["source_type"],
                "action": "create_new",
                "draft": proposal["draft"],
            }
        )

    imported = client.post(
        f"{API_V1}/ci/import/mapped",
        json={"items": items_payload, "type_mappings": mappings},
        headers=auth_headers,
    )
    assert imported.status_code == 200, imported.text
    body = report(imported.json())
    assert body["created"] == 3
    assert body["errors"] == []

    types = {t["name"]: t for t in items(client.get(f"{API_V1}/ci/types", headers=auth_headers).json())}
    assert type_a in types
    assert type_b in types
    assert types[type_a]["is_official"] is False

    for name in ci_names:
        listed = items(client.get(f"{API_V1}/ci?name={name}", headers=auth_headers).json())
        assert any(ci_row["name"] == name for ci_row in listed)

    created_ids = []
    for name in ci_names:
        listed = items(client.get(f"{API_V1}/ci?name={name}", headers=auth_headers).json())
        ci_row = next(c for c in listed if c["name"] == name)
        created_ids.append(ci_row["id"])

    for ci_id in created_ids:
        dr = client.delete(f"{API_V1}/ci/{ci_id}", headers=auth_headers)
        assert dr.status_code == 200, dr.text

    for type_name in (type_a, type_b):
        type_id = types[type_name]["id"]
        tr = client.delete(f"{API_V1}/ci/types/{type_id}", headers=auth_headers)
        assert tr.status_code == 200, tr.text


def test_v1_ci_handler_endpoints(client: TestClient, auth_headers: dict):
    """Cover v1 CI routes delegated to shared inventory handlers."""
    create = client.post(
        f"{API_V1}/ci",
        json={
            "name": "v1-handler-ci",
            "type_name": "Server",
            "owner": "ops",
            "environment": "test",
            "criticality": "low",
        },
        headers=auth_headers,
    )
    assert create.status_code == 200
    ci_id = ci(create.json())["id"]

    detail = client.get(f"{API_V1}/ci/{ci_id}", headers=auth_headers)
    assert detail.status_code == 200

    relations = client.get(f"{API_V1}/ci/{ci_id}/relations", headers=auth_headers)
    assert relations.status_code == 200

    deleted = client.delete(f"{API_V1}/ci/{ci_id}", headers=auth_headers)
    assert deleted.status_code == 200

    restored = client.post(f"{API_V1}/ci/{ci_id}/restore", headers=auth_headers)
    assert restored.status_code == 200

    deleted_again = client.delete(f"{API_V1}/ci/{ci_id}", headers=auth_headers)
    assert deleted_again.status_code == 200

    purged = client.delete(f"{API_V1}/ci/{ci_id}/purge", headers=auth_headers)
    assert purged.status_code == 200

    recycle = client.get(f"{API_V1}/ci/recycle-bin", headers=auth_headers)
    assert recycle.status_code == 200

    bulk = client.post(
        f"{API_V1}/ci/bulk/status",
        json={"ci_ids": [], "status": "active"},
        headers=auth_headers,
    )
    assert bulk.status_code == 200


def test_v1_relations_handler_update(client: TestClient, auth_headers: dict):
    db_id, app_id, _biz_id = _seed_chain(client, auth_headers)
    listed = client.get(f"{API_V1}/relations", headers=auth_headers)
    assert listed.status_code == 200
    rel = next(r for r in items(listed.json()) if r["source_ci_id"] == app_id and r["target_ci_id"] == db_id)
    updated = client.patch(
        f"{API_V1}/relations/{rel['id']}",
        json={"status": "active", "data_source": "manual"},
        headers=auth_headers,
    )
    assert updated.status_code == 200

    validated = client.get(f"{API_V1}/relations/validate", headers=auth_headers)
    assert validated.status_code == 200
    assert "valid" in validation(validated.json())

    deleted = client.delete(f"{API_V1}/relations/{rel['id']}", headers=auth_headers)
    assert deleted.status_code == 200
