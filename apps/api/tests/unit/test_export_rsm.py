from src.services.export.rsm import build_rsm_csv_zip, build_rsm_xlsx


def test_build_rsm_xlsx_two_sheets(db_session):
    from src.models import CI, CIType, Relation

    t = db_session.query(CIType).filter(CIType.name == "Server").first()
    a = CI(name="exp-a", type_id=t.id, status="active")
    b = CI(name="exp-b", type_id=t.id, status="active")
    db_session.add_all([a, b])
    db_session.flush()
    rel = Relation(source_ci_id=a.id, target_ci_id=b.id, relation_type="depends_on", status="active")
    db_session.add(rel)
    db_session.commit()

    buf = build_rsm_xlsx([a, b], [rel])
    from openpyxl import load_workbook

    wb = load_workbook(buf)
    assert wb.sheetnames == ["Elements", "Relations"]
    assert wb["Relations"].max_row == 2


def test_build_rsm_csv_zip_manifest(db_session):
    from src.models import CI, CIType

    t = db_session.query(CIType).filter(CIType.name == "Server").first()
    ci = CI(name="exp-only", type_id=t.id, status="active")
    db_session.add(ci)
    db_session.commit()

    buf = build_rsm_csv_zip([ci], [])
    import json
    import zipfile

    with zipfile.ZipFile(buf) as zf:
        manifest = json.loads(zf.read("manifest.json"))
        assert manifest["element_count"] == 1
        assert manifest["relation_count"] == 0
